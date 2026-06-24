#!/usr/bin/env python3
"""Convert the team's manual 'Transport Schedules' Excel into geocoded JSON fixtures.

Real-world gaps this handles:
  - Pallets are FRACTIONAL (3.5, 9.8, 0.4) -> kept as floats.
  - Addresses are FREE TEXT with no coordinates -> geocoded via a locality gazetteer
    (longest-substring match; production swaps this for a real geocoding API).
  - 'Vehicle' column actually holds the assigned TEAM (e.g. 'VMS Packers T3'); a vendor
    company can have several teams. 'Small Vehicle' in the name => 10ft (4 pallets).
  - 'Remarks' column holds the ORDER TYPE (Pickup / Retrieval / Partial / Intercity...).
  - 'Porter' / 'Porter Packers' are third-party on-demand aggregators => on_request tier.
"""
import json, sys, re
from pathlib import Path
import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else "/Users/safestorage/Downloads/21_06_2026 Transport Schedules Sheet.xlsx"
OUT = Path(__file__).resolve().parent.parent / "lib" / "fixtures"
DATE = "2026-06-21"

# locality -> (lat, lng). Approximate centroids; good enough to demonstrate routing.
GAZ = {
  # Bangalore
  "malur": (12.999, 77.938), "devarachikkanahalli": (12.886, 77.616), "btm": (12.916, 77.610),
  "marathahalli": (12.956, 77.701), "outer ring road": (12.956, 77.701), "peenya": (13.029, 77.519),
  "whitefield": (12.969, 77.749), "hrbr": (13.024, 77.643), "kalyan nagar": (13.024, 77.643),
  "electronic city": (12.845, 77.660), "hilalige": (12.84, 77.66), "cox town": (12.997, 77.617),
  "fraser town": (12.997, 77.617), "yelahanka": (13.100, 77.596), "doddaballapur": (13.13, 77.58),
  "kiadb": (13.13, 77.66), "gummanahalli": (13.13, 77.66), "wilson garden": (12.948, 77.598),
  "lakkasandra": (12.948, 77.598), "halanayakanahalli": (12.901, 77.687), "sarjapur": (12.901, 77.687),
  "kadubeesanahalli": (12.942, 77.693), "thubarahalli": (12.957, 77.717), "munnekolala": (12.957, 77.717),
  "halasuru": (12.978, 77.621), "ulsoor": (12.978, 77.621), "mahadevapura": (12.991, 77.687),
  "sonnenahalli": (12.979, 77.700), "doddanakundi": (12.979, 77.700), "kr puram": (13.007, 77.700),
  "narayanapura": (13.007, 77.700), "wheeler road": (12.997, 77.620), "cleveland town": (12.997, 77.620),
  "arehalli": (12.906, 77.546), "kodipur": (12.906, 77.546), "uttarahalli": (12.906, 77.546),
  # Hyderabad
  "puppalaguda": (17.412, 78.366), "kondapur": (17.4615, 78.364), "mahendra hills": (17.45, 78.53),
  "malkajgiri": (17.45, 78.53), "secunderabad": (17.44, 78.50), "rohini": (17.49, 78.52),
  "r k puram": (17.49, 78.52), "vidya nagar": (17.40, 78.51), "ram nagar": (17.40, 78.51),
  "medchal": (17.63, 78.48), "miyapur": (17.495, 78.358), "hafeezpet": (17.48, 78.36),
  "ameenpur": (17.52, 78.32), "gopanpalle": (17.48, 78.30), "tellapur": (17.48, 78.30),
}

CITY = {
  "BLR": {"city": "Bangalore", "center": (12.97, 77.59),
          "wh": {"lat": 12.991, "lng": 77.687, "label": "SafeStorage WH · Mahadevapura"}},
  "HYD": {"city": "Hyderabad", "center": (17.43, 78.42),
          "wh": {"lat": 17.46, "lng": 78.364, "label": "SafeStorage WH · Kondapur"}},
}

def geocode(addr, center):
    a = (addr or "").lower()
    best = None
    for key, coord in GAZ.items():
        if key in a and (best is None or len(key) > len(best[0])):
            best = (key, coord)
    if best:
        return best[1][0], best[1][1], best[0], True
    return center[0], center[1], None, False

def headers(ws):
    h = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v: h[str(v).strip().lower()] = c
    return h

def col(h, *names):
    for n in names:
        if n in h: return h[n]
    return None

def convert(sheet, meta):
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb[sheet]
    h = headers(ws)
    ci = col(h, "customer id"); cn = col(h, "customer name"); ca = col(h, "address")
    cp = col(h, "pallet"); cv = col(h, "vehicle"); cr = col(h, "remarks")
    cc = col(h, "transport charges"); cnotes = col(h, "customer notes")
    rows = []
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(r, ci).value
        if not cid or not str(cid).strip(): continue
        addr = ws.cell(r, ca).value
        if not addr: continue
        lat, lng, loc, hit = geocode(str(addr), meta["center"])
        team_raw = (ws.cell(r, cv).value or "").strip() if cv else ""
        typ = (ws.cell(r, cr).value or "").strip() if cr else ""
        pal = ws.cell(r, cp).value if cp else None
        try: pal = float(pal)
        except (TypeError, ValueError): pal = None
        ch = ws.cell(r, cc).value if cc else None
        try: ch = float(ch)
        except (TypeError, ValueError): ch = None
        # team_raw may be a status not a team
        status_words = ("reschedul", "cancel", "pending", "hold")
        is_status = any(w in team_raw.lower() for w in status_words)
        rows.append({
            "refNo": str(cid).strip(),
            "customerName": str(ws.cell(r, cn).value or "").strip() if cn else "",
            "address": str(addr).strip(),
            "locality": loc, "geocoded": hit,
            "lat": round(lat, 5), "lng": round(lng, 5),
            "pallets": pal, "type": typ or "Pickup",
            "team": "" if is_status else team_raw,
            "teamNote": team_raw if is_status else "",
            "charges": ch,
            "notes": str(ws.cell(r, cnotes).value or "").strip() if cnotes else "",
        })
    out = {"date": DATE, "city": meta["city"], "warehouse": meta["wh"], "bookings": rows}
    geocoded = sum(1 for x in rows if x["geocoded"])
    path = OUT / f"real-{sheet.lower()}.json"
    path.write_text(json.dumps(out, indent=2, ensure_ascii=False))
    print(f"{sheet}: {len(rows)} bookings, {geocoded} geocoded by locality -> {path.name}")

for sheet, meta in CITY.items():
    convert(sheet, meta)
